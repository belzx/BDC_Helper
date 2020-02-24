# 正式权重值
import os
import random
import logging

from openpyxl import Workbook, load_workbook

# 具体权重计算规则可以参考get_weight方法
# 下面一些参数比较重要
# 正式权重值（当hit_nums没有达到最大时，该值越大，则nums越大的单词出现的概率越大，更加大）
FORMAL_WEIGHT = 30
# 原始权重值（当hit_nums达到最大时，该值越大，则nums越大的单词出现的概率越大，更加大，如果设置为0，则该单词不会再出现）
ORIGIN_WIGHT = 10
# 负数权重值（当hit_nums小于0时，该值越大，则nums越大的单词出现的概率越大，更加大）
MINUS_WIGHT = 200
# 最大分值，要求大于0
MAX_HIT_NUM = 20
# 最小分值，要求小于0
MIN_HIT_NUM = -20
# 单词保证出现概率为0
NO_SHOW_HIT_NUM = -999
# tip
TIP_MESSAGE = "提示输入@,不知道#,不再显示$,上一个不再显示%,退出输入&"
# 正确得分
RIGHT_SCORE = 5
# 有提示后的得分
HELF_RIGHT_SCORE = 1
# 错误即不知道的得分
WRONG_SCORE = -7
# 问题类型占重
# 填充
FILL_QUESTION = 30
# 选择题
OPTION_QUESTION = 70
OPTION_EMUS = {
    0: "A",
    1: "B",
    2: "C",
    3: "D",
}
SCRIPT_PATH = os.path.abspath(os.path.dirname(__file__)).split("bdc.py")[0]
XL_PATH = os.path.join(SCRIPT_PATH, "dc_src.xlsx")

logging.basicConfig(level=logging.INFO,  # 控制台打印的日志级别
                    filename='bdc.log',
                    filemode='a',  ##模式，有w和a，w就是写模式，每次都会重新写日志，覆盖之前的日志
                    # a是追加模式，默认如果不写的话，就是追加模式
                    format=
                    '%(asctime)s : %(message)s'  # 日志格式

                    )
FLUSH_CACHE_SIZE = 10


class Word:
    def __init__(self, sheet_name, id, word, mean, nums, hit_nums, weight):
        """
        :param sheet_name:单词所在sheetName
        :param id: id
        :param word: 单词
        :param mean: 含义
        :param nums: 频率，频率越高，则随机挑选的这个单词的概率越高。
        :param hit_nums: 命中次数，具体使用看get weight方法
        :param weight: 权重，根据此权重决定随机挑选率
        """
        self.sheet_name = sheet_name
        self.id = id
        self.word = word
        self.mean = mean
        self.nums = nums
        self.hit_nums = hit_nums
        self.weight = weight


class Words:
    def __init__(self):
        self.content = {}
        # 权值的和，用来做随机取值
        self.sum = 0


class WordCollection:
    def __init__(self, excel: Workbook):
        self.name_content = {}
        self.excel = excel
        self.sheet_names = excel.sheetnames
        # id：flush list
        self.flush_list = {}

    def chose_sheet(self, sheet_name):
        self.sheet_name = sheet_name

    def add(self, word: Word):
        if word.sheet_name in self.name_content:
            words = self.name_content[word.sheet_name]
        else:
            words = Words()
            self.name_content[word.sheet_name] = words
        words.content[word.id] = word
        words.sum += word.weight

    def hint_one(self, word: Word, score: int):
        words = self.name_content[word.sheet_name]
        old_hit_nums = word.hit_nums
        # 重设命中值
        word.hit_nums = NO_SHOW_HIT_NUM if score == NO_SHOW_HIT_NUM else word.hit_nums + score
        if word.hit_nums == NO_SHOW_HIT_NUM:
            pass
        elif word.hit_nums > MAX_HIT_NUM:
            word.hit_nums = MAX_HIT_NUM
        elif word.hit_nums < MIN_HIT_NUM:
            word.hit_nums = MIN_HIT_NUM

        if old_hit_nums == word.hit_nums:
            return
        new_weight = get_weight(word)

        if word.weight == new_weight:
            pass
        else:
            words.sum = words.sum - word.weight + new_weight
            word.weight = new_weight
        # 刷入缓存重
        self.flush_list[word.id] = word.hit_nums
        if len(self.flush_list) >= FLUSH_CACHE_SIZE:
            self.excel_flush()

    def excel_flush(self):
        print_in_red("缓存开始刷新到磁盘")
        # 缓存刷新到硬盘
        for k in self.flush_list:
            table = self.excel[self.sheet_name]
            table.cell(row=k + 1, column=5).value = self.flush_list[k]
        self.excel.save(XL_PATH)
        print_in_red("刷新完毕")
        self.flush_list = {}

    def pick_one(self) -> Word:
        # 随机挑选，按权重
        words = self.name_content[self.sheet_name]
        random_int = random.randint(1, words.sum)
        sum = 0
        for a in words.content:
            sum += words.content[a].weight
            if sum >= random_int:
                return words.content[a]

    def pick_three_no_weight(self) -> list:
        # 随机挑选，不按照权重
        words = self.name_content[self.sheet_name]
        result = []
        for i in range(3):
            random_int = random.randint(1, len(words.content))
            result.append((words.content[random_int]))
        return result


def get_weight(word: Word) -> int:
    # 获取权重值
    if word.hit_nums == NO_SHOW_HIT_NUM:
        # 出现概率为0
        return 0
    elif word.hit_nums >= MAX_HIT_NUM:
        return word.hit_nums * ORIGIN_WIGHT
    elif word.hit_nums <= 0:
        return word.nums * FORMAL_WEIGHT + abs(word.hit_nums * MINUS_WIGHT)
    else:
        return word.nums * FORMAL_WEIGHT


def load() -> WordCollection:
    # 解析xlsx
    excel = load_workbook(XL_PATH)
    context = WordCollection(excel)
    for sheet_name in excel.sheetnames:
        # 加载信息到内存
        table = excel[sheet_name]
        rows = table.max_row
        for row in range(2, rows + 1):
            id = table.cell(row=row, column=1).value
            if not id:
                # 防止出现空行
                continue
            word = table.cell(row=row, column=2).value
            mean = table.cell(row=row, column=3).value
            nums = int(table.cell(row=row, column=4).value)
            hint_nums = int(table.cell(row=row, column=5).value)
            word = Word(sheet_name, id, word, mean, nums, hint_nums, 0)
            # 计算每个单词的权重
            word.weight = get_weight(word)
            # 加入到上下文
            context.add(word)
    return context


def init():
    context = load()
    message = "请输入数字，选择模块\n"
    for i in range(len(context.sheet_names)):
        message += "'%s => %s'" % (str(i), context.sheet_names[i])
    print_in_red(message)
    in_ = input()
    context.chose_sheet(context.sheet_names[int(in_)])
    # 上一个单词
    pre_word = None
    while True:
        one = context.pick_one()
        question = get_question_and_answer(context, one)
        print_in_wihte(question[0])
        have_tip = False
        while True:
            in_ = input().lower()
            print_in_green("input:[%s]" % in_)
            if in_ == "@":
                print_in_tip(question[1])
                have_tip = True
                pass
            elif in_ == "#":
                print_in_red(one.word)
                context.hint_one(one, WRONG_SCORE)
                print_in_red("不认识，得分[%s]" % WRONG_SCORE)
                break
            elif in_ == "$":
                # 后续不再显示此单词
                context.hint_one(one, NO_SHOW_HIT_NUM)
                print_in_red("后续不再显示此单词")
                break
            elif in_ == "%":
                # 后续不再显示上一个单词
                if pre_word:
                    context.hint_one(pre_word, NO_SHOW_HIT_NUM)
                print_in_red("后续不再显示上一个单词")
                continue
            elif in_ == "&":
                context.excel_flush()
                print_in_red("结束！！")
                return
            elif in_ == question[1]:
                if have_tip:
                    score = HELF_RIGHT_SCORE
                else:
                    score = RIGHT_SCORE
                context.hint_one(one, score)
                print_in_red("输入正确，得分[%s]" % str(score))
                break
            else:
                print_in_red("输入答案[%s]错误，请重新输入！" % in_)
                pass
        pre_word = one
    context.excel_flush()


def get_question_and_answer(context: WordCollection, word: Word) -> list:
    randomnum = random.randint(0, FILL_QUESTION + OPTION_QUESTION)
    message = ""
    right_answer = None
    if randomnum <= FILL_QUESTION:
        # 填空题
        message = "请输入（），该单词含义为%s[%s]" % (word.mean, TIP_MESSAGE,)
        right_answer = str(word.word).lower()
    else:
        # 选择题
        no_random = context.pick_three_no_weight()
        no_random.append(None)
        index = random.randint(0, 3)
        no_random.insert(index, word)
        message += "单词[%s]请选择正确的含义[%s]" % (word.word, TIP_MESSAGE)
        for i in range(4):
            message += "\n %s:%s" % (OPTION_EMUS[i], no_random[i].mean)
        right_answer = OPTION_EMUS[index].lower()

    return [message, right_answer]


def print_in_red(str):
    logging.info(str)
    print('\033[31m%s\033[0m' % str)


def print_in_green(str):
    logging.info(str)
    # print('\033[32m%s\033[0m' % str)


def print_in_tip(str):
    print_in_red("提示：%s,请重新输入！" % str)


def print_in_yellow(str):
    logging.info(str)
    # print('\033[33m%s\033[0m' % str)


def print_in_wihte(str):
    logging.info(str)
    print('\033[0m%s\033[0m' % str)


if __name__ == '__main__':
    os.system("")
    print(SCRIPT_PATH)
    print_in_red("START!!")
    init()
