# 正式权重值
import random
import time

from openpyxl import Workbook, load_workbook

from src import media_play
from src.constants import get_conf_center, XL_PATH, TIP_MESSAGE, OPTION_EMUS
from src.frequency_statistics import get_frequency_center
from src.utils import print_in_red, print_in_wihte, print_in_green, print_in_tip, is_number


class Word:
    def __init__(self, sheet_name, id, word, mean, nums, hit_nums, weight):
        """
        :param sheet_name:单词所在sheetName
        :param id: id
        :param word: 单词
        :param mean: 含义
        :param nums: 频率，频率越高，则随机挑选的这个单词的概率越高。
        :param hit_nums: 命中次数，具体使用看get weight方法
        :param weight: 权重，根据此权重决定随机挑选率
        """
        self.sheet_name = sheet_name
        self.id = id
        self.word = word
        self.mean = mean
        self.nums = nums
        self.hit_nums = hit_nums
        self.weight = weight


class Words:
    def __init__(self):
        self.content = {}
        # 权值的和，用来做随机取值
        self.sum = 0


class WordCollection:
    def __init__(self, excel: Workbook):
        self.name_content = {}
        self.excel = excel
        self.sheet_names = excel.sheetnames
        # id：flush list
        self.flush_list = {}

    def chose_sheet(self, sheet_name):
        self.sheet_name = sheet_name

    def add(self, word: Word):
        if word.sheet_name in self.name_content:
            words = self.name_content[word.sheet_name]
        else:
            words = Words()
            self.name_content[word.sheet_name] = words
        words.content[word.id] = word
        words.sum += word.weight

    def hint_one(self, word: Word, score: int):
        center = get_conf_center()

        # 更新频率设置
        if score > 0:
            if score == center.HALF_RIGHT_SCORE:
                f_score = center.FREQUENCY_HALF_SCORE
            else:
                f_score = center.FREQUENCY_SCORE
            get_frequency_center().update_status(word.sheet_name, word.id, word.word, f_score)

        words = self.name_content[word.sheet_name]
        old_hit_nums = word.hit_nums
        # 重设命中值
        word.hit_nums = center.NO_SHOW_HIT_NUM if score == center.NO_SHOW_HIT_NUM else word.hit_nums + score
        if word.hit_nums == center.NO_SHOW_HIT_NUM:
            pass
        elif word.hit_nums > center.MAX_HIT_NUM:
            word.hit_nums = center.MAX_HIT_NUM
        elif word.hit_nums < center.MIN_HIT_NUM:
            word.hit_nums = center.MIN_HIT_NUM

        if old_hit_nums == word.hit_nums:
            return
        new_weight = get_weight(word)

        if word.weight == new_weight:
            pass
        else:
            words.sum = words.sum - word.weight + new_weight
            word.weight = new_weight
        # 刷入缓存重
        self.flush_list[word.id] = word.hit_nums
        if len(self.flush_list) >= center.FLUSH_CACHE_SIZE:
            self.excel_flush()

    def excel_flush(self):
        print_in_red("缓存开始刷新到磁盘!")
        try:
            # 缓存刷新到硬盘
            for k in self.flush_list:
                table = self.excel[self.sheet_name]
                table.cell(row=k + 1, column=5).value = self.flush_list[k]
            self.excel.save(XL_PATH)
            print_in_red("刷新完毕!")
            self.flush_list = {}
        except Exception as e:
            print_in_red("写入失败，可能是excel被占用!")

    def pick_one(self) -> Word:
        # 随机挑选，按权重
        words = self.name_content[self.sheet_name]
        random_int = random.randint(1, int(words.sum))
        sum = 0
        for a in words.content:
            sum += words.content[a].weight
            if sum >= random_int:
                return words.content[a]

    def pick_three_no_weight(self) -> list:
        # 随机挑选，不按照权重
        words = self.name_content[self.sheet_name]
        result = []
        for i in range(3):
            random_int = random.randint(1, len(words.content))
            result.append((words.content[random_int]))
        return result

    def reload_words(self):
        # 重新加载权值
        sheet = self.name_content[self.sheet_name]
        sheet.sum = 0
        content = sheet.content
        for key in content:
            word = content[key]
            word.weight = get_weight(word)
            self.add(word)


def get_weight(word: Word) -> float:
    """"
    权重计算公式
    当hit_hums = 0 时
    weight = FORMAL_WEIGHT * word.nums

    当 max_hit_num > hit_hums > 0
    weight = （FORMAL_WEIGHT * word.nums + abs(ORIGIN_WEIGHT * word.hit_nums） * POSITIVE_WEIGHT

    当 min_hit_num < hit_hums < 0 时
    weight = （FORMAL_WEIGHT * word.nums + abs(ORIGIN_WEIGHT * word.hit_nums） * MINUS_WEIGHT

    当 hit_hums = max_hit_num 时
    weight = （FORMAL_WEIGHT * word.nums + abs(ORIGIN_WEIGHT * word.hit_nums） * MAX_POSITIVE_WEIGHT + MAX_POSITIVE_CONSTANT_WEIGHT

    当 hit_hums = min_hit_num 时
    weight = （FORMAL_WEIGHT * word.nums + abs(ORIGIN_WEIGHT * word.hit_nums） * MIN_MINUS_WEIGHT + MIN_MINUS_CONSTANT_WEIGHT
    """
    center = get_conf_center()
    if center.FREQUENCY_STATISTICS and not get_frequency_center().can_use(word.sheet_name, word.id):
        return 0

    # 获取权重值
    if word.hit_nums == center.NO_SHOW_HIT_NUM:
        # 出现概率为0
        return 0
    elif word.hit_nums == 0:
        return center.FORMAL_WEIGHT * word.nums
    elif center.MAX_HIT_NUM > word.hit_nums > 0:
        return (center.FORMAL_WEIGHT * word.nums + abs(center.ORIGIN_WEIGHT * word.hit_nums)) * center.POSITIVE_WEIGHT
    elif center.MIN_HIT_NUM < word.hit_nums < 0:
        return (center.FORMAL_WEIGHT * word.nums + abs(center.ORIGIN_WEIGHT * word.hit_nums)) * center.MINUS_WEIGHT
    elif center.MAX_HIT_NUM <= word.hit_nums:
        return (center.FORMAL_WEIGHT * word.nums + abs(
            center.ORIGIN_WEIGHT * word.hit_nums)) * center.MAX_POSITIVE_WEIGHT + center.MAX_POSITIVE_CONSTANT_WEIGHT
    elif center.MIN_HIT_NUM >= word.hit_nums:
        return (center.FORMAL_WEIGHT * word.nums + abs(
            center.ORIGIN_WEIGHT * word.hit_nums)) * center.MIN_MINUS_WEIGHT + center.MIN_MINUS_CONSTANT_WEIGHT
    else:
        print_in_red("计算权重出现错误!")
        return 0


def load() -> WordCollection:
    print_in_red("开始加载数据库!")
    # 解析xlsx
    excel = load_workbook(XL_PATH)
    context = WordCollection(excel)
    for sheet_name in excel.sheetnames:
        # 加载频率统计
        get_frequency_center().load_sheet(sheet_name)
        # 加载信息到内存
        table = excel[sheet_name]
        rows = table.max_row
        for row in range(2, rows + 1):
            id = table.cell(row=row, column=1).value
            if not id:
                # 防止出现空行
                continue
            word = table.cell(row=row, column=2).value
            mean = table.cell(row=row, column=3).value
            nums = float(table.cell(row=row, column=4).value)
            hit_nums = float(table.cell(row=row, column=5).value)
            word = Word(sheet_name, id, word, mean, nums, hit_nums, 0)
            # 计算每个单词的权重
            word.weight = get_weight(word)
            # 加入到上下文
            context.add(word)
    print_in_red("数据库加载完毕!")
    return context


def sound_play_before_input(question_type, have_tip, word):
    center = get_conf_center()
    if question_type in (Question.TYPE_MEANING, Question.TYPE_OPTION):
        media_play.play(center.PRONOUNCE_TYPE, word)
    elif question_type in (Question.TYPE_FILL,) and have_tip:
        # 默写还有提示才有发音
        media_play.play(center.PRONOUNCE_TYPE, word)


def sound_play_after_unknow(question_type, have_tip, word):
    center = get_conf_center()
    if question_type in (Question.TYPE_FILL,) and not have_tip:
        # 默写还有提示才有发音
        media_play.play(center.PRONOUNCE_TYPE, word)
        time.sleep(1.5)


def init():
    context = load()
    message = "请输入数字，选择模块\n"
    for i in range(len(context.sheet_names)):
        message += "'%s => %s'" % (str(i), context.sheet_names[i])
    print_in_red(message)
    in_ = input().lower().rstrip(" ").lstrip(" ")
    if not is_number(in_) or int(in_) > len(context.sheet_names) or int(in_) < 0:
        print_in_red("输入有误,程序终止!")
        return
    context.chose_sheet(context.sheet_names[int(in_)])
    # 上一个单词
    pre_word = None
    while True:
        one = context.pick_one()
        question = get_question_and_answer(context, one)
        print_in_wihte(question.message)
        have_tip = False
        while True:
            center = get_conf_center()
            #  判断是否发音,填空题和选择题才发音
            sound_play_before_input(question.type, have_tip, one.word)
            if center.HAD_UPDATE:
                # 重新计算权值
                print_in_red("配置有变动，重新修改权值")
                context.reload_words()
                center.HAD_UPDATE = False
                print_in_red("修改权值完毕！")
            in_ = input().lower().rstrip(" ").lstrip(" ")
            print_in_green("input:[%s]" % in_)
            if in_ == "@":
                print_in_tip(question.get_answer())
                have_tip = True
                pass
            elif in_ == "#":
                sound_play_after_unknow(question.type, have_tip, one.word)
                print_in_red(one.word)
                context.hint_one(one, center.WRONG_SCORE)
                print_in_red("不认识,正确答案[%s]，得分[%s]" % (question.get_answer(), center.WRONG_SCORE))
                break
            elif in_ == "$":
                # 后续不再显示此单词
                context.hint_one(one, center.NO_SHOW_HIT_NUM)
                print_in_red("后续不再显示此单词")
                break
            elif in_ == "%":
                # 后续不再显示上一个单词
                if pre_word:
                    context.hint_one(pre_word, center.NO_SHOW_HIT_NUM)
                print_in_red("后续不再显示上一个单词")
                continue
            elif in_ == "&":
                context.excel_flush()
                get_frequency_center().excel_flush()
                return
            elif in_ == "^":
                print_in_red("请输入要翻译的单词!")
                in_ = input().lower().rstrip(" ").lstrip(" ")
                print_in_green("input:[%s]" % in_)
                from src.youdao_dict import get_meaning
                print_in_red(get_meaning(in_))
                continue
            elif question.right(in_):
                if have_tip:
                    score = center.HALF_RIGHT_SCORE
                else:
                    score = center.RIGHT_SCORE
                context.hint_one(one, score)
                print_in_red("输入正确，得分[%s]" % str(score))
                break
            else:
                print_in_red("输入答案[%s]错误，请重新输入!" % in_)
                pass
        pre_word = one


QUESTION_INDEX = 0


class Question:
    TYPE_FILL = 1
    TYPE_OPTION = 2
    TYPE_MEANING = 3

    def __init__(self, type, message: str, answer):
        self.type = type
        self.message = message
        self.answer = answer

    def right(self, put_in) -> bool:
        if self.type in (Question.TYPE_FILL, Question.TYPE_OPTION):
            return self.answer == put_in
        elif self.type in (Question.TYPE_MEANING,):
            return put_in in self.answer

    def get_answer(self) -> str:
        if self.type in (Question.TYPE_FILL, Question.TYPE_OPTION):
            return self.answer
        elif self.type in (Question.TYPE_MEANING,):
            return " , ".join(self.answer)


def get_question_and_answer(context: WordCollection, word: Word) -> Question:
    global QUESTION_INDEX
    QUESTION_INDEX += 1
    center = get_conf_center()
    randomnum = random.randint(0, center.FILL_QUESTION + center.OPTION_QUESTION + center.MEANING_WRITE_QUESTION)
    message = ""
    if randomnum <= center.FILL_QUESTION:
        # 填空题
        type = Question.TYPE_FILL
        message = "请输入（），该单词含义为%s[%s]" % (word.mean, TIP_MESSAGE,)
        right_answer = str(word.word).lower()
    elif randomnum <= center.FILL_QUESTION + center.OPTION_QUESTION:
        # 选择题
        type = Question.TYPE_OPTION
        no_random = context.pick_three_no_weight()
        no_random.append(None)
        index = random.randint(0, 3)
        no_random.insert(index, word)
        message += "单词[%s]请选择正确的含义[%s]" % (word.word, TIP_MESSAGE)
        for i in range(4):
            message += "\n %s:%s" % (OPTION_EMUS[i], no_random[i].mean)
        right_answer = OPTION_EMUS[index].lower()
    else:
        # 默写题目
        type = Question.TYPE_MEANING
        message = "请输入写出[%s]的中文意思[%s]" % (word.word, TIP_MESSAGE,)
        right_answer = []
        for a in word.mean.split(","):
            right_answer.append(a.rstrip(" ").lstrip(" "))
    message = "[NO:%s][分值:%s]%s" % (QUESTION_INDEX, word.hit_nums, message)
    return Question(type, message, right_answer)
