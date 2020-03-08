import os
import threading
from time import sleep

from openpyxl import load_workbook
from pygame import mixer

import src.constants as constants
from src import utils
from src.constants import EN_PATH, AM_PATH, SCRIPT_PATH, MP3_TYPE_EN, get_conf_center
from src.utils import print_in_red

mixer.init()

XL_PATH = os.path.join(SCRIPT_PATH, "data", "dc_src.xlsx")
EN_URL = 'http://dict.youdao.com/dictvoice?type=1&audio=%s'  # 英式
AM_URL = 'http://dict.youdao.com/dictvoice?type=0&audio=%s'  # 美式


def download_mp3(file_url, save_url, file_name):
    res = utils.get_url_content(file_url, constants.get_conf_center().get_proxies())
    if res:
        # 解码
        file_path = os.path.join(save_url, file_name)
        if os.path.exists(file_path):
            print("[%s]已经存在" % file_path)
            return
        try:
            with open(file_path, 'wb') as fd:
                fd.write(res)
        except Exception as e:
            print(e)
            print("[%s]程序错误" % file_path)
            os.remove(file_path)


def download_by_word(word):
    download_mp3(AM_URL % word, AM_PATH, "%s.mp3" % word)
    download_mp3(EN_URL % word, EN_PATH, "%s.mp3" % word)


def download_by_wordlist(words: list):
    for word in words:
        download_by_word(word)


def play(type, word):
    center = get_conf_center()
    if not center.PRONOUNCE_SWITCH == "on":
        return
    if type == MP3_TYPE_EN:
        file_path = os.path.join(EN_PATH, "%s.mp3" % word)
    else:
        file_path = os.path.join(AM_PATH, "%s.mp3" % word)

    # 先判断文件是否存在
    if not os.path.exists(file_path):
        if type == MP3_TYPE_EN:
            download_mp3(EN_URL % word, EN_PATH, "%s.mp3" % word)
        else:
            download_mp3(AM_URL % word, AM_PATH, "%s.mp3" % word)
    try:
        mixer.music.load(file_path)
        mixer.music.play()
    except Exception as a:
        print_in_red(a)


def download_all_words():
    excel = load_workbook(XL_PATH)
    words = set()
    for sheet_name in excel.sheetnames:
        table = excel[sheet_name]
        rows = table.max_row
        for row in range(2, rows + 1):
            id = table.cell(row=row, column=1).value
            if not id:
                # 防止出现空行
                continue
            words.add(table.cell(row=row, column=2).value)

    lists = []
    index = 0
    size = 250
    words = list(words)
    while True:
        l = words[index:index + size]
        if l:
            lists.append(l)
            index += size
        else:
            break

    tlist = []
    for ll in lists:
        t = threading.Thread(target=download_by_wordlist, args=(ll,), name="Thread-A")
        tlist.append(t)
        t.start()

    for t in tlist:
        t.join()

    print("Over")


if __name__ == '__main__':
    play(2, "A")
    sleep(2)
