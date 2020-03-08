import datetime
import os

import openpyxl
import requests
from openpyxl import Workbook, load_workbook

from src.constants import FREQUENCY_PATH, get_conf_center
from src.utils import print_in_red

STATUS_TO_PERIOD = {
    1: 1,
    2: 1,
    3: 2,
    4: 3,
    5: 8,
    6: 15,
}

MAX_STATUS = len(STATUS_TO_PERIOD)


class FrequencyWord:
    def __init__(self, sheet_name, id, word, status, score, latest_time: datetime):
        self.sheet_name = sheet_name
        self.id = id
        self.word = word
        self.status = status
        self.score = score
        self.latest_time = latest_time


class FrequencyCenter:
    def __init__(self):
        self.excel = self.get_excel()
        self.sheet_name = None
        self.content = {}
        self.flush_list = {}

    def load_sheet(self, sheet_name):
        if not sheet_name in self.content:
            self.content[sheet_name] = {}
        if not sheet_name in self.excel.sheetnames:
            print_in_red("创建新的sheet[%s]" % sheet_name)
            self.excel.create_sheet(sheet_name)
            table = self.excel[sheet_name]
            table.cell(row=1, column=1).value = "id"
            table.cell(row=1, column=2).value = "word"
            table.cell(row=1, column=3).value = "status"
            table.cell(row=1, column=4).value = "score"
            table.cell(row=1, column=5).value = "latest_update_time"
            self.excel.save(FREQUENCY_PATH)
        else:
            table = self.excel[sheet_name]
            rows = table.max_row
            for row in range(2, rows + 1):
                id = table.cell(row=row, column=1).value
                if not id:
                    continue
                word = table.cell(row=row, column=2).value
                status = table.cell(row=row, column=3).value
                score = table.cell(row=row, column=4).value
                time = str_to_time(table.cell(row=row, column=5).value)
                frequency_word = FrequencyWord(sheet_name, id, word, status, score, time)
                self.add(frequency_word)

    def add(self, word: FrequencyWord):
        lo = self.content[word.sheet_name]
        lo[word.id] = word

    def can_use(self, sheet_name, id) -> bool:
        center = get_conf_center()
        if not sheet_name in self.content:
            self.load_sheet(sheet_name)
        if not center.FREQUENCY_STATISTICS:
            return True
        name_ = self.content[sheet_name]
        if not id in name_:
            return True
        fword = name_[id]
        if fword == 0:  # 为0 一个阶段才开始
            return (datetime.datetime.now() - fword.latest_time).days >= STATUS_TO_PERIOD[fword.status]
        else:
            return True

    def update_status(self, sheet_name, id, word, score):
        if not sheet_name in self.content:
            self.load_sheet(sheet_name)
        center = get_conf_center()
        if not center.FREQUENCY_STATISTICS:
            return
        if score <= 0:
            return
        if not id in self.content[sheet_name]:
            if score >= center.MAX_FREQUENCY_SCORE:
                fword = FrequencyWord(sheet_name, id, word, 2, 0, datetime.datetime.now())
            else:
                fword = FrequencyWord(sheet_name, id, word, 1, score, datetime.datetime.now())
            self.content[sheet_name][id] = fword
        else:
            fword = self.content[sheet_name][id]
            fword.score += score
            if fword.status >= MAX_STATUS:
                pass
            else:
                if fword.score >= center.MAX_FREQUENCY_SCORE:
                    fword.score = 0
                    fword.status += 1
                    fword.latest_time = datetime.datetime.now()
        self.flush_list[id] = fword
        if len(self.flush_list) >= center.MAX_FREQUENCY_FLUSH_SIZE:
            self.excel_flush()

    def get_excel(self) -> Workbook:
        if not os.path.exists(FREQUENCY_PATH):
            wb = openpyxl.Workbook()
            wb.save(FREQUENCY_PATH)
        return load_workbook(FREQUENCY_PATH)

    def excel_flush(self):
        center = get_conf_center()
        if not center.FREQUENCY_STATISTICS:
            return
        try:
            print_in_red("开始刷入频率统计缓存!")
            for k in self.flush_list:
                fword = self.flush_list[k]
                table = self.excel[fword.sheet_name]
                table.cell(row=k + 1, column=1).value = fword.id
                table.cell(row=k + 1, column=2).value = fword.word
                table.cell(row=k + 1, column=3).value = fword.status
                table.cell(row=k + 1, column=4).value = fword.score
                table.cell(row=k + 1, column=5).value = time_to_str(fword.latest_time)
            self.excel.save(FREQUENCY_PATH)
            self.flush_list = {}
            print_in_red("刷新完毕!")
        except Exception as e:
            print_in_red(e)
            print_in_red("写入失败,可能是excel被占用!")


def time_to_str(strftime: datetime) -> str:
    return datetime.datetime.strftime(strftime, "%Y-%m-%d %H:%M:%S")


def str_to_time(strftime: str) -> datetime:
    return datetime.datetime.strptime(strftime, "%Y-%m-%d %H:%M:%S")


FREQUENCY_CENTER = FrequencyCenter()

def get_frequency_center() -> FrequencyCenter:
    return FREQUENCY_CENTER
